"""
CAPEX / OPEX lookup module.

Loads ReLIFE_CAPEX.xlsx and ReLIFE_OPEX.xlsx once at module import time and
exposes compute_capex() / compute_opex() for the risk-assessment service.
"""
from __future__ import annotations

import statistics
from pathlib import Path
from typing import Protocol, runtime_checkable

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_DATA_DIR = Path(__file__).parent
_CAPEX_FILE = _DATA_DIR / "ReLIFE_CAPEX.xlsx"
_OPEX_FILE = _DATA_DIR / "ReLIFE_OPEX.xlsx"

# ---------------------------------------------------------------------------
# Action classification
# ---------------------------------------------------------------------------
_AREA_ACTIONS: frozenset[str] = frozenset({
    "Wall insulation",
    "Wall insulation - additional",
    "Roof insulation - Accessible",
    "Roof insulation - Makeover",
    "Floor insulation",
    "Windows",
})
_CAPACITY_ACTIONS: frozenset[str] = frozenset({
    "Air-water Heat Pump",
    "Condensing boiler",
    "PV",
    "Solar thermal panels",
})

ALL_ACTIONS: list[str] = sorted(_AREA_ACTIONS | _CAPACITY_ACTIONS)

# OPEX column header → canonical renovation action name
_OPEX_COLUMN_MAP: dict[str, str] = {
    "Heat Pump":     "Air-water Heat Pump",
    "Gas Boiler":    "Condensing boiler",
    "PV":            "PV",
    "Solar Thermal": "Solar thermal panels",
}

# ---------------------------------------------------------------------------
# Country normalisation
# ---------------------------------------------------------------------------

def _normalise_country(name: str) -> str:
    """Lower-case and strip; map 'czechia' → 'czech republic'."""
    n = name.strip().lower()
    if n == "czechia":
        n = "czech republic"
    return n


# ---------------------------------------------------------------------------
# Internal storage
# ---------------------------------------------------------------------------
_CapexKey = tuple[str, str]  # (country_norm, action)

# capex_table[(country_norm, action)] = list of row-dicts
_capex_table: dict[_CapexKey, list[dict]] = {}
_capex_countries: set[str] = set()

# opex_table[country_norm][action] = annual_cost (EUR/year)
_opex_table: dict[str, dict[str, float]] = {}
_opex_countries: set[str] = set()


# ---------------------------------------------------------------------------
# Loaders (called once at module import)
# ---------------------------------------------------------------------------

def _load_capex() -> None:
    wb = openpyxl.load_workbook(_CAPEX_FILE, data_only=True, read_only=True)
    ws = wb.active
    first_row = True
    for row in ws.iter_rows(values_only=True):
        if first_row:
            first_row = False
            continue
        country_raw = row[1]
        action = row[3]
        if not country_raw or not action:
            continue
        country_norm = _normalise_country(str(country_raw))
        _capex_countries.add(country_norm)
        material = row[4]
        price_m2 = row[11]    # Price [€/m²]
        price_fixed = row[12]  # Price [€]  (fixed installation cost)
        price_kw = row[13]     # Price [€/kW]
        entry: dict = {
            "material":    str(material) if material is not None else None,
            "price_m2":    float(price_m2) if price_m2 is not None else None,
            "price_fixed": float(price_fixed) if price_fixed is not None else None,
            "price_kw":    float(price_kw) if price_kw is not None else None,
        }
        key: _CapexKey = (country_norm, action)
        _capex_table.setdefault(key, []).append(entry)
    wb.close()


def _load_opex() -> None:
    wb = openpyxl.load_workbook(_OPEX_FILE, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    # Map column index → canonical action name
    col_action: dict[int, str] = {
        i: _OPEX_COLUMN_MAP[h]
        for i, h in enumerate(headers)
        if h in _OPEX_COLUMN_MAP
    }
    for row in rows[1:]:
        if not row[0]:
            continue
        country_norm = _normalise_country(str(row[0]))
        _opex_countries.add(country_norm)
        _opex_table[country_norm] = {
            action_name: float(row[col_idx]) if row[col_idx] is not None else 0.0
            for col_idx, action_name in col_action.items()
        }
    wb.close()


# Eagerly load both files on module import
_load_capex()
_load_opex()


# ---------------------------------------------------------------------------
# Action protocol — structural typing to avoid circular imports with models
# ---------------------------------------------------------------------------

@runtime_checkable
class ActionLike(Protocol):
    action: str
    area_m2: float | None
    capacity_kw: float | None
    material: str | None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def supported_countries() -> list[str]:
    """Return the 27 supported EU country names (canonical spelling)."""
    return sorted({
        "Austria", "Belgium", "Bulgaria", "Croatia", "Cyprus",
        "Czech Republic", "Denmark", "Estonia", "Finland", "France",
        "Germany", "Greece", "Hungary", "Ireland", "Italy",
        "Latvia", "Lithuania", "Luxembourg", "Malta", "Netherlands",
        "Poland", "Portugal", "Romania", "Slovakia", "Slovenia",
        "Spain", "Sweden",
    })


def supported_actions() -> list[str]:
    """Return the 10 supported renovation action names."""
    return ALL_ACTIONS


def compute_capex(country: str, actions: list) -> float:
    """
    Compute total CAPEX (EUR) for a renovation package in the given country.

    Parameters
    ----------
    country:
        EU country name (case-insensitive). Both "Czech Republic" and "Czechia"
        are accepted.
    actions:
        List of action-like objects exposing ``action``, ``area_m2``,
        ``capacity_kw``, and ``material`` attributes.

    Raises
    ------
    ValueError
        If the country is unsupported, an action name is unrecognised, or a
        required quantity (area_m2 / capacity_kw) is missing.
    """
    country_norm = _normalise_country(country)
    if country_norm not in _capex_countries:
        raise ValueError(
            f"Country '{country}' is not supported. "
            f"Supported: {', '.join(supported_countries())}"
        )

    total = 0.0
    for act in actions:
        action_name: str = act.action
        if action_name not in _AREA_ACTIONS and action_name not in _CAPACITY_ACTIONS:
            raise ValueError(
                f"Unknown renovation action '{action_name}'. "
                f"Supported: {', '.join(ALL_ACTIONS)}"
            )

        rows = _capex_table.get((country_norm, action_name))
        if not rows:
            raise ValueError(
                f"No CAPEX data for country='{country}', action='{action_name}'."
            )

        candidates = _filter_by_material(rows, act.material)

        if action_name in _AREA_ACTIONS:
            if act.area_m2 is None:
                raise ValueError(f"area_m2 is required for action '{action_name}'.")
            prices = [r["price_m2"] for r in candidates if r["price_m2"] is not None]
            if not prices:
                raise ValueError(
                    f"No price_m2 data for country='{country}', action='{action_name}'."
                )
            total += statistics.mean(prices) * act.area_m2

        else:  # capacity-based
            if act.capacity_kw is None:
                raise ValueError(f"capacity_kw is required for action '{action_name}'.")
            fixed_prices = [r["price_fixed"] for r in candidates if r["price_fixed"] is not None]
            kw_prices = [r["price_kw"] for r in candidates if r["price_kw"] is not None]
            price_fixed = statistics.mean(fixed_prices) if fixed_prices else 0.0
            price_kw = statistics.mean(kw_prices) if kw_prices else 0.0
            total += price_fixed + price_kw * act.capacity_kw

    return total


def compute_opex(country: str, actions: list) -> float:
    """
    Compute total annual O&M cost (EUR/year) for a renovation package.

    Insulation and windows have no OPEX entry and contribute 0.

    Raises
    ------
    ValueError
        If the country is not supported.
    """
    country_norm = _normalise_country(country)
    if country_norm not in _opex_countries:
        raise ValueError(
            f"Country '{country}' is not supported. "
            f"Supported: {', '.join(supported_countries())}"
        )

    country_opex = _opex_table.get(country_norm, {})
    return sum(country_opex.get(act.action, 0.0) for act in actions)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _filter_by_material(rows: list[dict], material: str | None) -> list[dict]:
    """
    Return rows matching the requested material (case-insensitive).
    If material is None or no rows match, return all rows for averaging.
    """
    if material is None:
        return rows
    matched = [r for r in rows if r["material"] and r["material"].lower() == material.lower()]
    return matched if matched else rows
